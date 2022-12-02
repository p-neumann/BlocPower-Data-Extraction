from openpyxl import Workbook

class WorksheetWriter:

    def __init__(self):
        self.wb = Workbook()
        self.workSheet = None
        self.createWorksheet()

    def createWorksheet(self) -> None:
        self.workSheet = self.wb.create_sheet("Building Data")

    def writeDataToWorksheetRow(self, data, row, sheet) -> None:
        
        for i in range(len(data)):
            sheet.cell(row=row, column=i + 1).value = data[i]
    
    def writeDataToWorksheet(self, dataList) -> None:

        sheet = self.wb.active
        row = 1
        for data in dataList:
            self.writeDataToWorksheetRow(data, row, sheet)
            row += 1
        self.wb.save("extractedData.xlsx")